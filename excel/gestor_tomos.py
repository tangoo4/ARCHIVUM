"""Operaciones de lectura y modificación de tomos ya guardados."""

import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


FILA_INICIO_DATOS = 2
COLUMNAS_TOMO = "ABCDEFGH"

FILL_ROJO = PatternFill("solid", fgColor="FF3030")
FILL_AMARILLO = PatternFill("solid", fgColor="FFE600")
FILL_BLANCO = PatternFill("solid", fgColor="FFFFFF")
ALINEACION_CENTRADA = Alignment(horizontal="center", vertical="center")


def leer_tomos(ruta_excel):
    """Lee únicamente las filas que representan tomos reales."""
    wb = load_workbook(Path(ruta_excel), data_only=False)
    ws = wb.active
    tomos = []

    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        try:
            tomo = int(ws[f"A{fila}"].value)
        except (TypeError, ValueError):
            continue

        tomos.append({
            "fila": fila,
            "tomo": tomo,
            "anio": ws[f"B{fila}"].value,
            "matriz_inicio": ws[f"C{fila}"].value,
            "fecha_inicio": ws[f"D{fila}"].value,
            "matriz_final": ws[f"E{fila}"].value,
            "fecha_final": ws[f"F{fila}"].value,
            "medida": ws[f"G{fila}"].value,
            "observaciones": ws[f"H{fila}"].value or "",
        })

    return tomos


def comprobar_continuidad(tomos, indice, matriz_inicio, matriz_final):
    """Devuelve avisos si la corrección rompe la continuidad con otro tomo."""
    avisos = []
    if indice > 0:
        try:
            esperado = int(tomos[indice - 1]["matriz_final"]) + 1
            if matriz_inicio != esperado:
                avisos.append(f"La matriz inicial esperada según el tomo anterior es {esperado}.")
        except (TypeError, ValueError):
            pass

    if indice < len(tomos) - 1:
        try:
            esperado_siguiente = matriz_final + 1
            inicio_siguiente = int(tomos[indice + 1]["matriz_inicio"])
            if inicio_siguiente != esperado_siguiente:
                avisos.append(
                    f"El tomo siguiente empieza en {inicio_siguiente}; debería empezar en {esperado_siguiente}."
                )
        except (TypeError, ValueError):
            pass
    return avisos


def modificar_tomo(ruta_excel, tomo, datos, medida_estandar):
    """Actualiza A:H en la fila del tomo sin alterar foliado ni columnas auxiliares."""
    ruta_excel = Path(ruta_excel)
    wb = load_workbook(ruta_excel, data_only=False)
    ws = wb.active

    fila_objetivo = None
    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        try:
            if int(ws[f"A{fila}"].value) == int(tomo):
                fila_objetivo = fila
                break
        except (TypeError, ValueError):
            continue

    if fila_objetivo is None:
        raise ValueError(f"No existe el tomo {tomo} en la temporada.")

    valores = (
        int(tomo),
        datos["anio"],
        int(datos["matriz_inicio"]),
        datos["fecha_inicio"],
        int(datos["matriz_final"]),
        datos["fecha_final"],
        datos["medida"],
        datos["observaciones"],
    )
    for columna, valor in zip(COLUMNAS_TOMO, valores):
        ws[f"{columna}{fila_objetivo}"] = valor

    medida = datos["medida"]
    if medida == "?":
        fill = FILL_AMARILLO
    elif medida != medida_estandar:
        fill = FILL_ROJO
    else:
        fill = FILL_BLANCO

    for columna in COLUMNAS_TOMO:
        celda = ws[f"{columna}{fila_objetivo}"]
        celda.fill = copy.copy(fill)
        celda.alignment = copy.copy(ALINEACION_CENTRADA)

    wb.save(ruta_excel)

