"""Gestor Excel v0.3"""
from pathlib import Path
from openpyxl import load_workbook
from shutil import copy2

class GestorExcel:
    def __init__(self, plantilla):
        self.plantilla=Path(plantilla)
        self.libro=None
        self.hoja=None
        self.archivo=None

    def crear_temporada(self,destino):
        destino=Path(destino)
        copy2(self.plantilla,destino)
        self.abrir(destino)
        return destino

    def abrir(self,archivo):
        self.archivo=Path(archivo)
        self.libro=load_workbook(self.archivo)
        self.hoja=self.libro[self.libro.sheetnames[0]]

    def buscar_siguiente_fila(self):
        fila=2
        while self.hoja[f"A{fila}"].value not in (None,""):
            fila+=1
        return fila

    def ultimo_tomo(self):
        return self.buscar_siguiente_fila()-2

    def escribir_tomo(self,tomo,anio,mi,fi,mf,ff,medida,obs):
        fila=self.buscar_siguiente_fila()
        self.hoja[f"A{fila}"]=tomo
        self.hoja[f"B{fila}"]=anio
        self.hoja[f"C{fila}"]=mi
        self.hoja[f"D{fila}"]=fi
        self.hoja[f"E{fila}"]=mf
        self.hoja[f"F{fila}"]=ff
        self.hoja[f"G{fila}"]=medida
        self.hoja[f"H{fila}"]=obs
        self.libro.save(self.archivo)
        return fila

    def leer_tabla(self):
        datos=[]
        fila=2
        while self.hoja[f"A{fila}"].value not in (None,""):
            datos.append([self.hoja[f"{c}{fila}"].value for c in "ABCDEFGH"])
            fila+=1
        return datos
