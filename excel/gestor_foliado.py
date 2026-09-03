"""Lectura y escritura del foliado de una temporada."""

from pathlib import Path

from openpyxl import load_workbook


FILA_INICIO_DATOS = 2
COL_TOMO = "A"
COL_FOLIADO = "J"
COL_HOJAS = "K"


def leer_foliado(ruta_excel):
    """Devuelve los tomos reales y su foliado, ordenados como aparecen en Excel."""
    ruta_excel = Path(ruta_excel)
    wb = load_workbook(ruta_excel, data_only=False)
    ws = wb.active
    tomos = []

    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        try:
            tomo = int(ws[f"{COL_TOMO}{fila}"].value)
        except (TypeError, ValueError):
            continue

        foliado = ws[f"{COL_FOLIADO}{fila}"].value
        hojas = ws[f"{COL_HOJAS}{fila}"].value
        tomos.append({"fila": fila, "tomo": tomo, "foliado": foliado, "hojas": hojas})

    return tomos


def guardar_foliado(ruta_excel, tomo, foliado):
    """Guarda J y recalcula K desde el tomo modificado hasta el último foliado."""
    ruta_excel = Path(ruta_excel)
    wb = load_workbook(ruta_excel, data_only=False)
    ws = wb.active

    filas = []
    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        try:
            numero_tomo = int(ws[f"{COL_TOMO}{fila}"].value)
        except (TypeError, ValueError):
            continue
        filas.append((fila, numero_tomo))

    indice = next((i for i, (_, numero) in enumerate(filas) if numero == tomo), None)
    if indice is None:
        raise ValueError(f"No existe el tomo {tomo} en la temporada.")

    foliado = int(foliado)
    if foliado <= 0:
        raise ValueError("El foliado debe ser un número mayor que cero.")

    if indice > 0:
        fila_anterior, _ = filas[indice - 1]
        foliado_anterior = ws[f"{COL_FOLIADO}{fila_anterior}"].value
        if foliado_anterior in (None, ""):
            raise ValueError("Primero debes introducir el foliado del tomo anterior.")
        if foliado <= int(foliado_anterior):
            raise ValueError(f"El foliado debe ser mayor que {int(foliado_anterior)}.")

    if indice < len(filas) - 1:
        fila_siguiente, tomo_siguiente = filas[indice + 1]
        foliado_siguiente = ws[f"{COL_FOLIADO}{fila_siguiente}"].value
        if foliado_siguiente not in (None, "") and foliado >= int(foliado_siguiente):
            raise ValueError(
                f"El foliado debe ser menor que {int(foliado_siguiente)}, "
                f"que ya está guardado para el tomo {tomo_siguiente}."
            )

    fila_actual, _ = filas[indice]
    ws[f"{COL_FOLIADO}{fila_actual}"] = foliado

    # Guardamos valores, no fórmulas, para que el resultado sea visible incluso
    # antes de que Excel o LibreOffice recalcule el libro.
    for posicion in range(indice, len(filas)):
        fila, _ = filas[posicion]
        valor_actual = ws[f"{COL_FOLIADO}{fila}"].value
        if valor_actual in (None, ""):
            ws[f"{COL_HOJAS}{fila}"] = None
            continue

        valor_actual = int(valor_actual)
        if posicion == 0:
            hojas = valor_actual
        else:
            fila_previa, _ = filas[posicion - 1]
            valor_previo = ws[f"{COL_FOLIADO}{fila_previa}"].value
            if valor_previo in (None, ""):
                ws[f"{COL_HOJAS}{fila}"] = None
                continue
            hojas = valor_actual - int(valor_previo)
            if hojas <= 0:
                ws[f"{COL_HOJAS}{fila}"] = None
                continue

        ws[f"{COL_HOJAS}{fila}"] = hojas

    wb.save(ruta_excel)
