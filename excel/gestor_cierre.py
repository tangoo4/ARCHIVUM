"""Operaciones para marcar el cierre del medido de una temporada."""

import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

FILA_INICIO_DATOS = 2
COL_TOMO = "A"
COL_CIERRE_INICIO = "C"
COL_CIERRE_FIN = "D"
FILL_CIERRE = PatternFill("solid", fgColor="FFE600")


def buscar_ultima_fila_tomo(ws):
    """Devuelve la fila del último tomo real o None si no existen tomos."""
    ultima_fila = None
    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        try:
            int(ws[f"{COL_TOMO}{fila}"].value)
        except (TypeError, ValueError):
            continue
        ultima_fila = fila
    return ultima_fila


def _rango_cierre(fila):
    return f"{COL_CIERRE_INICIO}{fila}:{COL_CIERRE_FIN}{fila}"


def temporada_cerrada(ruta_excel):
    """Indica si existe una línea de cierre justo debajo del último tomo."""
    wb = load_workbook(ruta_excel)
    ws = wb.active
    ultima_fila = buscar_ultima_fila_tomo(ws)
    if ultima_fila is None:
        return False

    fila_cierre = ultima_fila + 1
    rango = _rango_cierre(fila_cierre)
    return rango in {str(r) for r in ws.merged_cells.ranges}


def cerrar_temporada(ruta_excel, notario):
    """Crea la línea de cierre y guarda el Excel.

    Devuelve la fila utilizada para el cierre.
    """
    wb = load_workbook(ruta_excel)
    ws = wb.active

    ultima_fila = buscar_ultima_fila_tomo(ws)
    if ultima_fila is None:
        raise ValueError("Todavía no existen tomos medidos.")

    fila_cierre = ultima_fila + 1
    rango = _rango_cierre(fila_cierre)
    rangos_combinados = {str(r) for r in ws.merged_cells.ranges}
    if rango in rangos_combinados:
        raise RuntimeError("Esta temporada ya contiene una línea de cierre.")

    borde_c = copy.copy(ws[f"C{ultima_fila}"].border)
    borde_d = copy.copy(ws[f"D{ultima_fila}"].border)
    alineacion = copy.copy(ws[f"C{ultima_fila}"].alignment)
    fuente = copy.copy(ws[f"C{ultima_fila}"].font)

    ws.merge_cells(rango)
    celda = ws[f"C{fila_cierre}"]
    celda.value = notario
    celda.fill = copy.copy(FILL_CIERRE)
    celda.border = borde_c
    celda.alignment = alineacion
    celda.font = fuente

    # Mantiene el borde exterior derecho del criterio de la fila anterior.
    ws[f"D{fila_cierre}"].fill = copy.copy(FILL_CIERRE)
    ws[f"D{fila_cierre}"].border = borde_d
    ws.row_dimensions[fila_cierre].height = ws.row_dimensions[ultima_fila].height

    wb.save(ruta_excel)
    return fila_cierre
