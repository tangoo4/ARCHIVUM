"""
==========================================================
ARCHIVUM
Ventana: Control
Versión: 0.6.1
==========================================================

Módulo de control de temporada.

Pestañas:
- Estado
- Validación
- Incidencias
- Historial
- Producción / Tapas fabricadas

Funcional:
- Estado
- Validación
- Incidencias
- Producción básica con tapas fabricadas guardadas en produccion.json
"""

import json
from pathlib import Path
from tkinter import messagebox
from tkinter import ttk

import customtkinter as ctk
from openpyxl import load_workbook

from config import (
    COLOR_BG,
    COLOR_PANEL,
    COLOR_PANEL_2,
    COLOR_BORDER,
    COLOR_TEXT,
    COLOR_TEXT_MUTED,
    COLOR_GREEN,
    COLOR_CYAN,
    COLOR_RED,
    COLOR_YELLOW,
    FONT_SUBTITLE,
    FONT_NORMAL,
)


FILA_INICIO_DATOS = 2

COL_TOMO = "A"
COL_ANY = "B"
COL_PROT_INICIAL = "C"
COL_DATA_INICIAL = "D"
COL_PROT_FINAL = "E"
COL_DATA_FINAL = "F"
COL_GRUIX = "G"
COL_OBSERVACIONS = "H"


class PanelControl(ctk.CTkFrame):
    """
    Área de Control integrada en el espacio de trabajo de la temporada.
    """

    def __init__(self, master, app):
        super().__init__(master, fg_color=COLOR_BG)

        self.app = app
        self.contexto = app.contexto
        self.ruta_excel = Path(self.contexto.archivo_actual) if self.contexto.archivo_actual else None
        self.carpeta_temporada = self.ruta_excel.parent if self.ruta_excel else None
        self.ruta_produccion = self.carpeta_temporada / "produccion.json" if self.carpeta_temporada else None
        self.ruta_historial = self.carpeta_temporada / "historial.json" if self.carpeta_temporada else None

        self.medida_estandar = self.contexto.medida_estandar or "8,5"
        self.filas = self._leer_excel()
        self.produccion = self._leer_produccion()

        self._crear_interfaz()

    # ======================================================
    # LECTURA EXCEL / JSON
    # ======================================================

    def _leer_excel(self):
        if not self.ruta_excel or not self.ruta_excel.exists():
            messagebox.showerror(
                "Archivo no encontrado",
                "No hay ningún Excel abierto para controlar.",
            )
            return []

        wb = load_workbook(self.ruta_excel, data_only=True)
        ws = wb.active

        filas = []

        for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
            tomo = ws[f"{COL_TOMO}{fila}"].value

            if tomo in (None, ""):
                continue

            filas.append({
                "fila_excel": fila,
                "tomo": ws[f"{COL_TOMO}{fila}"].value,
                "anio": ws[f"{COL_ANY}{fila}"].value,
                "matriz_inicio": ws[f"{COL_PROT_INICIAL}{fila}"].value,
                "fecha_inicio": ws[f"{COL_DATA_INICIAL}{fila}"].value,
                "matriz_final": ws[f"{COL_PROT_FINAL}{fila}"].value,
                "fecha_final": ws[f"{COL_DATA_FINAL}{fila}"].value,
                "medida": ws[f"{COL_GRUIX}{fila}"].value,
                "observaciones": ws[f"{COL_OBSERVACIONS}{fila}"].value or "",
            })

        return filas

    def _leer_produccion(self):
        """
        Lee produccion.json.

        Estructura:
        {
            "version": 1,
            "fabricadas": [1, 2, 3]
        }
        """

        if not self.ruta_produccion:
            return {"version": 1, "fabricadas": []}

        if not self.ruta_produccion.exists():
            data = {"version": 1, "fabricadas": []}
            self._guardar_produccion(data)
            return data

        try:
            with open(self.ruta_produccion, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = {"version": 1, "fabricadas": []}

        if "fabricadas" not in data:
            data["fabricadas"] = []

        return data

    def _guardar_produccion(self, data=None):
        if data is None:
            data = self.produccion

        if not self.ruta_produccion:
            return

        self.ruta_produccion.parent.mkdir(parents=True, exist_ok=True)

        with open(self.ruta_produccion, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

    # ======================================================
    # INTERFAZ GENERAL
    # ======================================================

    def _crear_interfaz(self):
        self.tabs = ctk.CTkTabview(
            self,
            fg_color=COLOR_PANEL,
            segmented_button_fg_color=COLOR_PANEL_2,
            segmented_button_selected_color=COLOR_GREEN,
            segmented_button_selected_hover_color=COLOR_GREEN,
            segmented_button_unselected_color=COLOR_PANEL_2,
            segmented_button_unselected_hover_color=COLOR_BORDER,
            text_color=COLOR_TEXT,
        )
        self.tabs.pack(fill="both", expand=True, padx=20, pady=(10, 20))

        self.tab_estado = self.tabs.add("Estado")
        self.tab_validacion = self.tabs.add("Validación")
        self.tab_incidencias = self.tabs.add("Incidencias")
        self.tab_historial = self.tabs.add("Historial")
        self.tab_produccion = self.tabs.add("Producción")

        self._crear_estado()
        self._crear_validacion()
        self._crear_incidencias()
        self._crear_historial()
        self._crear_produccion()

    def _crear_cabecera(self):
        cabecera = ctk.CTkFrame(self, fg_color=COLOR_PANEL, height=90, corner_radius=0)
        cabecera.pack(fill="x")

        ctk.CTkLabel(
            cabecera,
            text="CONTROL",
            font=("Segoe UI", 30, "bold"),
            text_color=COLOR_GREEN,
        ).place(x=25, y=15)

        archivo = str(self.ruta_excel) if self.ruta_excel else "SIN ARCHIVO"

        ctk.CTkLabel(
            cabecera,
            text=archivo,
            font=FONT_NORMAL,
            text_color=COLOR_TEXT_MUTED,
        ).place(x=27, y=57)

    def _crear_panel(self, master):
        panel = ctk.CTkFrame(
            master,
            fg_color=COLOR_PANEL_2,
            border_width=1,
            border_color=COLOR_BORDER,
            corner_radius=12,
        )
        panel.pack(fill="both", expand=True, padx=15, pady=15)
        return panel

    # ======================================================
    # ESTADO
    # ======================================================

    def _crear_estado(self):
        panel = self._crear_panel(self.tab_estado)

        datos = self._calcular_estado()

        ctk.CTkLabel(
            panel,
            text="ESTADO DE TEMPORADA",
            font=FONT_SUBTITLE,
            text_color=COLOR_CYAN,
        ).pack(anchor="w", padx=20, pady=(20, 10))

        tabla = ctk.CTkFrame(panel, fg_color="transparent")
        tabla.pack(anchor="nw", padx=30, pady=10)

        for i, (clave, valor) in enumerate(datos):
            ctk.CTkLabel(
                tabla,
                text=clave,
                font=FONT_NORMAL,
                text_color=COLOR_TEXT_MUTED,
                width=250,
                anchor="w",
            ).grid(row=i, column=0, sticky="w", pady=5)

            ctk.CTkLabel(
                tabla,
                text=str(valor),
                font=FONT_NORMAL,
                text_color=COLOR_TEXT,
                width=300,
                anchor="w",
            ).grid(row=i, column=1, sticky="w", pady=5)

    def _calcular_estado(self):
        if not self.filas:
            return [("Estado", "Sin datos")]

        tomos = [self._to_int(f["tomo"]) for f in self.filas if self._to_int(f["tomo"]) is not None]
        inicios = [self._to_int(f["matriz_inicio"]) for f in self.filas if self._to_int(f["matriz_inicio"]) is not None]
        finales = [self._to_int(f["matriz_final"]) for f in self.filas if self._to_int(f["matriz_final"]) is not None]

        medidas = [str(f["medida"] or "") for f in self.filas]
        observaciones = [str(f["observaciones"] or "").strip() for f in self.filas]

        estandar = sum(1 for m in medidas if m == self.medida_estandar)
        especiales = sum(1 for m in medidas if m not in ("", "?", self.medida_estandar))
        pendientes = sum(1 for m in medidas if m == "?")
        obs = sum(1 for o in observaciones if o)

        ultima_fecha = ""
        for f in reversed(self.filas):
            if f["fecha_final"] not in (None, ""):
                ultima_fecha = f["fecha_final"]
                break

        fabricadas = len(set(self.produccion.get("fabricadas", [])))
        total = len(self.filas)
        porcentaje = round((fabricadas / total) * 100, 1) if total else 0

        return [
            ("Tipo", self.contexto.tipo or "-"),
            ("Notario", self.contexto.notario or "-"),
            ("Año", self.contexto.anio or "-"),
            ("Tomos", total),
            ("Primer tomo", min(tomos) if tomos else "-"),
            ("Último tomo", max(tomos) if tomos else "-"),
            ("Primera matriz", min(inicios) if inicios else "-"),
            ("Última matriz", max(finales) if finales else "-"),
            ("Medida estándar", self.medida_estandar),
            ("Tomos estándar", estandar),
            ("Tomos especiales", especiales),
            ("Tomos con ?", pendientes),
            ("Observaciones", obs),
            ("Última fecha final", ultima_fecha or "-"),
            ("Tapas fabricadas", f"{fabricadas} / {total} ({porcentaje} %)"),
        ]

    # ======================================================
    # VALIDACIÓN
    # ======================================================

    def _crear_validacion(self):
        panel = self._crear_panel(self.tab_validacion)

        resultado, incidencias = self._validar_temporada()

        color = COLOR_GREEN if resultado == "CORRECTO" else COLOR_YELLOW if resultado == "REVISAR" else COLOR_RED

        ctk.CTkLabel(
            panel,
            text=f"RESULTADO: {resultado}",
            font=("Segoe UI", 24, "bold"),
            text_color=color,
        ).pack(anchor="w", padx=20, pady=(20, 10))

        contenedor = ctk.CTkFrame(panel, fg_color=COLOR_PANEL, corner_radius=10)
        contenedor.pack(fill="both", expand=True, padx=20, pady=(10, 20))

        self.tabla_validacion = ttk.Treeview(
            contenedor,
            columns=("tipo", "detalle"),
            show="headings",
        )

        self.tabla_validacion.heading("tipo", text="TIPO")
        self.tabla_validacion.heading("detalle", text="DETALLE")

        self.tabla_validacion.column("tipo", width=160, anchor="center")
        self.tabla_validacion.column("detalle", width=780, anchor="w")

        scroll = ttk.Scrollbar(contenedor, orient="vertical", command=self.tabla_validacion.yview)
        self.tabla_validacion.configure(yscrollcommand=scroll.set)

        self.tabla_validacion.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        self._estilo_treeview()
        self.tabla_validacion.tag_configure("ok", foreground=COLOR_GREEN)
        self.tabla_validacion.tag_configure("warn", foreground=COLOR_YELLOW)
        self.tabla_validacion.tag_configure("error", foreground=COLOR_RED)

        for tipo, detalle, tag in incidencias:
            self.tabla_validacion.insert("", "end", values=(tipo, detalle), tags=(tag,))

    def _validar_temporada(self):
        incidencias = []
        errores = 0
        avisos = 0

        if not self.filas:
            return "ERROR", [("ERROR", "No hay tomos para validar.", "error")]

        tomos = [self._to_int(f["tomo"]) for f in self.filas]
        tomos = [t for t in tomos if t is not None]

        if tomos:
            for esperado, real in enumerate(tomos, start=tomos[0]):
                if real != esperado:
                    errores += 1
                    incidencias.append(("ERROR", f"Tomo no consecutivo. Esperado {esperado}, encontrado {real}.", "error"))
                    break
            else:
                incidencias.append(("OK", "Tomos consecutivos.", "ok"))

        filas_ordenadas = sorted(self.filas, key=lambda f: self._to_int(f["tomo"]) or 0)

        matriz_error = False

        for i in range(1, len(filas_ordenadas)):
            anterior = filas_ordenadas[i - 1]
            actual = filas_ordenadas[i]

            final_anterior = self._to_int(anterior["matriz_final"])
            inicio_actual = self._to_int(actual["matriz_inicio"])

            if final_anterior is None or inicio_actual is None:
                continue

            esperado = final_anterior + 1

            if inicio_actual != esperado:
                matriz_error = True
                errores += 1
                incidencias.append((
                    "ERROR",
                    f"Tomo {actual['tomo']}: matriz inicial {inicio_actual}. Esperada {esperado}.",
                    "error",
                ))

        if not matriz_error:
            incidencias.append(("OK", "Matrices consecutivas.", "ok"))

        for f in self.filas:
            inicio = self._to_int(f["matriz_inicio"])
            final = self._to_int(f["matriz_final"])

            if inicio is not None and final is not None and final < inicio:
                errores += 1
                incidencias.append((
                    "ERROR",
                    f"Tomo {f['tomo']}: protocolo final menor que inicial.",
                    "error",
                ))

        pendientes = [f for f in self.filas if str(f["medida"] or "") == "?"]
        especiales = [
            f for f in self.filas
            if str(f["medida"] or "") not in ("", "?", self.medida_estandar)
        ]

        if pendientes:
            avisos += len(pendientes)
            incidencias.append(("AVISO", f"{len(pendientes)} tomos con medida pendiente (?).", "warn"))

        if especiales:
            avisos += len(especiales)
            incidencias.append(("AVISO", f"{len(especiales)} tomos con medida especial.", "warn"))

        if errores:
            return "ERROR", incidencias

        if avisos:
            return "REVISAR", incidencias

        return "CORRECTO", incidencias

    # ======================================================
    # INCIDENCIAS
    # ======================================================

    def _crear_incidencias(self):
        panel = self._crear_panel(self.tab_incidencias)

        izquierda = ctk.CTkFrame(panel, fg_color="transparent")
        izquierda.pack(side="left", fill="both", expand=True, padx=(15, 8), pady=15)

        derecha = ctk.CTkFrame(panel, fg_color="transparent")
        derecha.pack(side="left", fill="both", expand=True, padx=(8, 15), pady=15)

        ctk.CTkLabel(
            izquierda,
            text="OBSERVACIONES",
            font=FONT_SUBTITLE,
            text_color=COLOR_CYAN,
        ).pack(anchor="w", pady=(0, 10))

        self.tabla_observaciones = self._crear_tabla(
            izquierda,
            columnas=("tomo", "observacion"),
            encabezados={"tomo": "TOMO", "observacion": "OBSERVACIÓN"},
            anchos={"tomo": 80, "observacion": 360},
        )

        ctk.CTkLabel(
            derecha,
            text="MATRICES PENDIENTES",
            font=FONT_SUBTITLE,
            text_color=COLOR_CYAN,
        ).pack(anchor="w", pady=(0, 10))

        self.tabla_matrices = self._crear_tabla(
            derecha,
            columnas=("matriz",),
            encabezados={"matriz": "MATRIZ"},
            anchos={"matriz": 240},
        )

        botones = ctk.CTkFrame(derecha, fg_color="transparent")
        botones.pack(fill="x", pady=(10, 0))

        ctk.CTkButton(
            botones,
            text="COPIAR MATRICES",
            command=self._copiar_matrices,
            fg_color=COLOR_GREEN,
            text_color="black",
            hover_color=COLOR_GREEN,
            height=38,
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            botones,
            text="COPIAR OBSERVACIONES",
            command=self._copiar_observaciones,
            fg_color=COLOR_PANEL_2,
            text_color=COLOR_TEXT,
            hover_color=COLOR_BORDER,
            height=38,
        ).pack(side="left")

        self._cargar_incidencias()

    def _crear_tabla(self, master, columnas, encabezados, anchos):
        contenedor = ctk.CTkFrame(master, fg_color=COLOR_PANEL, corner_radius=10)
        contenedor.pack(fill="both", expand=True)

        tabla = ttk.Treeview(contenedor, columns=columnas, show="headings")

        for col in columnas:
            tabla.heading(col, text=encabezados[col])
            tabla.column(col, width=anchos[col], anchor="center")

        if "observacion" in columnas:
            tabla.column("observacion", anchor="w")

        scroll = ttk.Scrollbar(contenedor, orient="vertical", command=tabla.yview)
        tabla.configure(yscrollcommand=scroll.set)

        tabla.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        self._estilo_treeview()

        return tabla

    def _cargar_incidencias(self):
        matrices = set()

        for f in self.filas:
            obs = str(f["observaciones"] or "").strip()

            if not obs:
                continue

            self.tabla_observaciones.insert("", "end", values=(f["tomo"], obs))

            for numero in self._extraer_matrices(obs):
                matrices.add(numero)

        for matriz in sorted(matrices):
            self.tabla_matrices.insert("", "end", values=(matriz,))

    def _extraer_matrices(self, texto):
        partes = texto.split(",")
        numeros = []

        for parte in partes:
            parte = parte.strip()

            if parte.isdigit():
                numeros.append(int(parte))

        return numeros

    def _copiar_matrices(self):
        matrices = []

        for item in self.tabla_matrices.get_children():
            valor = self.tabla_matrices.item(item, "values")[0]
            matrices.append(str(valor))

        texto = "\n".join(matrices)

        self.clipboard_clear()
        self.clipboard_append(texto)

        messagebox.showinfo("Copiado", "Matrices copiadas al portapapeles.")

    def _copiar_observaciones(self):
        lineas = []

        for item in self.tabla_observaciones.get_children():
            tomo, obs = self.tabla_observaciones.item(item, "values")
            lineas.append(f"{tomo}\t{obs}")

        texto = "\n".join(lineas)

        self.clipboard_clear()
        self.clipboard_append(texto)

        messagebox.showinfo("Copiado", "Observaciones copiadas al portapapeles.")

    # ======================================================
    # HISTORIAL
    # ======================================================

    def _crear_historial(self):
        panel = self._crear_panel(self.tab_historial)

        ctk.CTkLabel(
            panel,
            text="HISTORIAL DE MODIFICACIONES",
            font=FONT_SUBTITLE,
            text_color=COLOR_CYAN,
        ).pack(anchor="w", padx=20, pady=(20, 10))

        ctk.CTkLabel(
            panel,
            text=(
                "Preparado para la siguiente fase.\n\n"
                "Aquí se mostrarán modificaciones posteriores sobre tomos ya existentes:\n"
                "- cambios de medida\n"
                "- cambios de observaciones\n"
                "- correcciones de matrices\n\n"
                "No registrará la creación normal de tomos para evitar ruido."
            ),
            font=FONT_NORMAL,
            text_color=COLOR_TEXT_MUTED,
            justify="left",
        ).pack(anchor="w", padx=25, pady=10)

    # ======================================================
    # PRODUCCIÓN
    # ======================================================

    def _crear_produccion(self):
        panel = self._crear_panel(self.tab_produccion)

        cabecera = ctk.CTkFrame(panel, fg_color="transparent")
        cabecera.pack(fill="x", padx=20, pady=(20, 10))

        ctk.CTkLabel(
            cabecera,
            text="PRODUCCIÓN - TAPAS FABRICADAS",
            font=FONT_SUBTITLE,
            text_color=COLOR_CYAN,
        ).pack(anchor="w")

        self.label_produccion = ctk.CTkLabel(
            cabecera,
            text="",
            font=("Segoe UI", 18, "bold"),
            text_color=COLOR_GREEN,
        )
        self.label_produccion.pack(anchor="w", pady=(10, 5))

        self.barra_produccion = ctk.CTkProgressBar(
            cabecera,
            width=500,
            height=18,
            progress_color=COLOR_GREEN,
            fg_color=COLOR_PANEL,
        )
        self.barra_produccion.pack(anchor="w", pady=(0, 10))

        contenedor = ctk.CTkFrame(panel, fg_color=COLOR_PANEL, corner_radius=10)
        contenedor.pack(fill="both", expand=True, padx=20, pady=(10, 20))

        self.tabla_produccion = ttk.Treeview(
            contenedor,
            columns=("tomo", "fabricada"),
            show="headings",
        )

        self.tabla_produccion.heading("tomo", text="TOMO")
        self.tabla_produccion.heading("fabricada", text="TAPA FABRICADA")

        self.tabla_produccion.column("tomo", width=140, anchor="center")
        self.tabla_produccion.column("fabricada", width=220, anchor="center")

        scroll = ttk.Scrollbar(contenedor, orient="vertical", command=self.tabla_produccion.yview)
        self.tabla_produccion.configure(yscrollcommand=scroll.set)

        self.tabla_produccion.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        self._estilo_treeview()

        self.tabla_produccion.bind("<Double-1>", self._toggle_tapa_evento)
        self.tabla_produccion.bind("<space>", self._toggle_tapa_evento)

        botones = ctk.CTkFrame(panel, fg_color="transparent")
        botones.pack(fill="x", padx=20, pady=(0, 15))

        ctk.CTkButton(
            botones,
            text="MARCAR / DESMARCAR",
            command=self._toggle_tapa_seleccionada,
            fg_color=COLOR_GREEN,
            hover_color=COLOR_GREEN,
            text_color="black",
            height=38,
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            botones,
            text="MARCAR TODAS",
            command=self._marcar_todas,
            fg_color=COLOR_PANEL_2,
            hover_color=COLOR_BORDER,
            text_color=COLOR_TEXT,
            height=38,
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            botones,
            text="DESMARCAR TODAS",
            command=self._desmarcar_todas,
            fg_color=COLOR_PANEL_2,
            hover_color=COLOR_BORDER,
            text_color=COLOR_TEXT,
            height=38,
        ).pack(side="left")

        self._cargar_produccion()

    def _cargar_produccion(self):
        for item in self.tabla_produccion.get_children():
            self.tabla_produccion.delete(item)

        fabricadas = set(self._to_int(t) for t in self.produccion.get("fabricadas", []))
        fabricadas.discard(None)

        tomos = sorted(
            [self._to_int(f["tomo"]) for f in self.filas if self._to_int(f["tomo"]) is not None]
        )

        for tomo in tomos:
            estado = "☑" if tomo in fabricadas else "☐"
            self.tabla_produccion.insert("", "end", iid=str(tomo), values=(tomo, estado))

        self._actualizar_resumen_produccion()

    def _actualizar_resumen_produccion(self):
        total = len(self.filas)
        fabricadas = len(set(self.produccion.get("fabricadas", [])))
        porcentaje = round((fabricadas / total) * 100, 1) if total else 0

        self.label_produccion.configure(
            text=f"{fabricadas} / {total} tapas fabricadas ({porcentaje} %)"
        )

        progreso = fabricadas / total if total else 0
        self.barra_produccion.set(progreso)

    def _toggle_tapa_evento(self, _event=None):
        self._toggle_tapa_seleccionada()

    def _toggle_tapa_seleccionada(self):
        seleccionado = self.tabla_produccion.selection()

        if not seleccionado:
            return

        tomo = self._to_int(seleccionado[0])

        if tomo is None:
            return

        fabricadas = set(self._to_int(t) for t in self.produccion.get("fabricadas", []))
        fabricadas.discard(None)

        if tomo in fabricadas:
            fabricadas.remove(tomo)
        else:
            fabricadas.add(tomo)

        self.produccion["fabricadas"] = sorted(fabricadas)
        self._guardar_produccion()
        self._cargar_produccion()

    def _marcar_todas(self):
        tomos = sorted(
            [self._to_int(f["tomo"]) for f in self.filas if self._to_int(f["tomo"]) is not None]
        )
        self.produccion["fabricadas"] = tomos
        self._guardar_produccion()
        self._cargar_produccion()

    def _desmarcar_todas(self):
        self.produccion["fabricadas"] = []
        self._guardar_produccion()
        self._cargar_produccion()

    # ======================================================
    # UTILIDADES
    # ======================================================

    def _to_int(self, valor):
        try:
            return int(valor)
        except Exception:
            return None

    def _estilo_treeview(self):
        style = ttk.Style()
        style.theme_use("default")

        style.configure(
            "Treeview",
            background="#181818",
            foreground=COLOR_TEXT,
            fieldbackground="#181818",
            rowheight=28,
            font=("Segoe UI", 12),
        )

        style.configure(
            "Treeview.Heading",
            background=COLOR_PANEL_2,
            foreground=COLOR_CYAN,
            font=("Segoe UI", 12, "bold"),
        )

        style.map("Treeview", background=[("selected", "#003C5A")])
