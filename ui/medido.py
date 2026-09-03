"""
==========================================================
ARCHIVUM
Pantalla: Medido
Versión: 0.7.1
==========================================================

Cambios v0.7.1:\n- Corregida detección de temporada nueva: el tomo 1 vuelve a tener matriz inicial libre.\n\nCambios v0.5.0:
- Botón BACKUP sustituido por CONTROL.
- CONTROL abre Estado, Validación, Incidencias, Historial y Producción.

Cambios previos:
- Autocompletado de fechas con día + mes: 1 EN -> 1 ENERO.
- Validación de días según mes: evita 40 ENERO o 31 FEBRERO.
- Mantiene protocolo final >= inicial, fecha inicial automática, colores, bordes y centrado.
"""

import copy
from pathlib import Path
from tkinter import messagebox
from tkinter import ttk

import customtkinter as ctk
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

from ui.control import PanelControl
from ui.foliado import VentanaFoliado
from ui.modificar_tomo import VentanaModificarTomo
from excel.gestor_cierre import cerrar_temporada, temporada_cerrada

from config import (
    COLOR_BG,
    COLOR_PANEL,
    COLOR_PANEL_2,
    COLOR_BORDER,
    COLOR_TEXT,
    COLOR_TEXT_MUTED,
    COLOR_GREEN,
    COLOR_CYAN,
    COLOR_RED,
    COLOR_YELLOW,
    FONT_SUBTITLE,
    FONT_NORMAL,
    FONT_TOMO,
    DEFAULT_MEDIDA,
    MAX_MEDIDA,
    MAX_OBSERVACIONES,
)


FILA_INICIO_DATOS = 2

COL_TOMO = "A"
COL_ANY = "B"
COL_PROT_INICIAL = "C"
COL_DATA_INICIAL = "D"
COL_PROT_FINAL = "E"
COL_DATA_FINAL = "F"
COL_GRUIX = "G"
COL_OBSERVACIONS = "H"

COLUMNAS_COLOR_FILA = [
    COL_TOMO,
    COL_ANY,
    COL_PROT_INICIAL,
    COL_DATA_INICIAL,
    COL_PROT_FINAL,
    COL_DATA_FINAL,
    COL_GRUIX,
    COL_OBSERVACIONS,
]

FILL_ROJO = PatternFill("solid", fgColor="FF3030")
FILL_AMARILLO = PatternFill("solid", fgColor="FFE600")
FILL_BLANCO = PatternFill("solid", fgColor="FFFFFF")

ALINEACION_CENTRADA = Alignment(horizontal="center", vertical="center")

MESES_AUTOCOMPLETAR = {
    "E": "ENERO", "EN": "ENERO", "ENE": "ENERO", "ENERO": "ENERO",
    "F": "FEBRERO", "FE": "FEBRERO", "FEB": "FEBRERO", "FEBRERO": "FEBRERO",
    "MAR": "MARZO", "MARZ": "MARZO", "MARZO": "MARZO",
    "AB": "ABRIL", "ABR": "ABRIL", "ABRI": "ABRIL", "ABRIL": "ABRIL",
    "MAY": "MAYO", "MAYO": "MAYO",
    "JUN": "JUNIO", "JUNI": "JUNIO", "JUNIO": "JUNIO",
    "JUL": "JULIO", "JULI": "JULIO", "JULIO": "JULIO",
    "AG": "AGOSTO", "AGO": "AGOSTO", "AGOS": "AGOSTO", "AGOSTO": "AGOSTO",
    "S": "SEPTIEMBRE", "SE": "SEPTIEMBRE", "SEP": "SEPTIEMBRE", "SEPT": "SEPTIEMBRE",
    "SET": "SEPTIEMBRE", "SEPTIEMBRE": "SEPTIEMBRE",
    "O": "OCTUBRE", "OC": "OCTUBRE", "OCT": "OCTUBRE", "OCTUBRE": "OCTUBRE",
    "N": "NOVIEMBRE", "NO": "NOVIEMBRE", "NOV": "NOVIEMBRE", "NOVIEMBRE": "NOVIEMBRE",
    "D": "DICIEMBRE", "DI": "DICIEMBRE", "DIC": "DICIEMBRE", "DICIEMBRE": "DICIEMBRE",
}

DIAS_MAXIMOS_MES = {
    "ENERO": 31,
    "FEBRERO": 29,
    "MARZO": 31,
    "ABRIL": 30,
    "MAYO": 31,
    "JUNIO": 30,
    "JULIO": 31,
    "AGOSTO": 31,
    "SEPTIEMBRE": 30,
    "OCTUBRE": 31,
    "NOVIEMBRE": 30,
    "DICIEMBRE": 31,
}


class PantallaMedido(ctk.CTkFrame):
    """Pantalla de trabajo del medido."""

    def __init__(self, master, app):
        super().__init__(master, fg_color=COLOR_BG)

        self.app = app
        self.contexto = app.contexto

        self.ruta_excel = Path(self.contexto.archivo_actual) if self.contexto.archivo_actual else None
        self.medida_estandar = self.contexto.medida_estandar or DEFAULT_MEDIDA

        self.tomo_actual = 1
        self.ultima_fila = FILA_INICIO_DATOS
        self.ultima_matriz_final = None
        self.ultima_fecha_final = None
        self.matriz_inicio_esperada = None

        self.pack(fill="both", expand=True)

        self._cargar_estado_excel()

        self._crear_interfaz()
        self._configurar_enter()
        self._actualizar_tomo()
        self._actualizar_matriz_esperada()
        self._refrescar_tabla_desde_excel()

    # ======================================================
    # EXCEL
    # ======================================================

    def _abrir_libro(self):
        if not self.ruta_excel or not self.ruta_excel.exists():
            raise FileNotFoundError(f"No existe el archivo Excel:\n{self.ruta_excel}")

        wb = load_workbook(self.ruta_excel)
        ws = wb.active
        return wb, ws

    def _cargar_estado_excel(self):
        if not self.ruta_excel or not self.ruta_excel.exists():
            self.tomo_actual = self.contexto.tomo_actual or 1
            self.ultima_fila = FILA_INICIO_DATOS
            self.ultima_matriz_final = None
            self.ultima_fecha_final = None
            self.matriz_inicio_esperada = None
            return

        wb, ws = self._abrir_libro()

        ultima_fila_con_datos = FILA_INICIO_DATOS - 1

        for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
            tomo = ws[f"{COL_TOMO}{fila}"].value

            # Solo consideramos una fila como tomo real si la columna TOMO
            # contiene un número. Esto evita que estilos, fórmulas o restos
            # de la plantilla hagan creer a Archivum que ya hay tomos guardados.
            try:
                int(tomo)
            except Exception:
                continue

            ultima_fila_con_datos = fila

        if ultima_fila_con_datos < FILA_INICIO_DATOS:
            self.tomo_actual = 1
            self.ultima_fila = FILA_INICIO_DATOS
            self.ultima_matriz_final = None
            self.ultima_fecha_final = None
            self.matriz_inicio_esperada = None
            return

        ultimo_tomo = ws[f"{COL_TOMO}{ultima_fila_con_datos}"].value
        ultima_matriz = ws[f"{COL_PROT_FINAL}{ultima_fila_con_datos}"].value
        ultima_fecha_final = ws[f"{COL_DATA_FINAL}{ultima_fila_con_datos}"].value

        try:
            self.tomo_actual = int(ultimo_tomo) + 1
        except Exception:
            self.tomo_actual = 1

        self.ultima_fila = ultima_fila_con_datos + 1

        try:
            self.ultima_matriz_final = int(ultima_matriz)
            self.matriz_inicio_esperada = self.ultima_matriz_final + 1
        except Exception:
            self.ultima_matriz_final = None
            self.ultima_fecha_final = None
            self.matriz_inicio_esperada = None

        self.ultima_fecha_final = str(ultima_fecha_final) if ultima_fecha_final not in (None, "") else None

        self.contexto.tomo_actual = self.tomo_actual

    def _guardar_en_excel(self, datos):
        wb, ws = self._abrir_libro()
        fila = self.ultima_fila

        self._copiar_estilo_base(ws, fila)

        ws[f"{COL_TOMO}{fila}"] = datos["tomo"]
        ws[f"{COL_ANY}{fila}"] = datos["anio"]
        ws[f"{COL_PROT_INICIAL}{fila}"] = int(datos["matriz_inicio"])
        ws[f"{COL_DATA_INICIAL}{fila}"] = datos["fecha_inicio"]
        ws[f"{COL_PROT_FINAL}{fila}"] = int(datos["matriz_final"])
        ws[f"{COL_DATA_FINAL}{fila}"] = datos["fecha_final"]
        ws[f"{COL_GRUIX}{fila}"] = datos["medida"]
        ws[f"{COL_OBSERVACIONS}{fila}"] = datos["observaciones"]

        self._aplicar_color_fila(ws, fila, datos["medida"])
        self._centrar_fila(ws, fila)

        wb.save(self.ruta_excel)

    def _copiar_estilo_base(self, ws, fila):
        """
        Copia bordes, fuente, formato y dimensiones desde una fila base.

        Si existe fila anterior, se usa la anterior.
        Si es el primer tomo, se usa la fila 2 de la plantilla.
        """

        if fila > FILA_INICIO_DATOS:
            fila_base = fila - 1
        else:
            fila_base = FILA_INICIO_DATOS

        for columna in COLUMNAS_COLOR_FILA:
            origen = ws[f"{columna}{fila_base}"]
            destino = ws[f"{columna}{fila}"]

            if origen.has_style:
                destino._style = copy.copy(origen._style)

            destino.font = copy.copy(origen.font)
            destino.border = copy.copy(origen.border)
            destino.number_format = origen.number_format
            destino.protection = copy.copy(origen.protection)

        ws.row_dimensions[fila].height = ws.row_dimensions[fila_base].height

    def _aplicar_color_fila(self, ws, fila, medida):
        if medida == "?":
            fill = FILL_AMARILLO
        elif medida != self.medida_estandar:
            fill = FILL_ROJO
        else:
            fill = FILL_BLANCO

        for columna in COLUMNAS_COLOR_FILA:
            ws[f"{columna}{fila}"].fill = copy.copy(fill)

    def _centrar_fila(self, ws, fila):
        for columna in COLUMNAS_COLOR_FILA:
            ws[f"{columna}{fila}"].alignment = copy.copy(ALINEACION_CENTRADA)

    def _leer_filas_excel(self):
        if not self.ruta_excel or not self.ruta_excel.exists():
            return []

        wb, ws = self._abrir_libro()
        filas = []

        for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
            tomo = ws[f"{COL_TOMO}{fila}"].value

            try:
                int(tomo)
            except Exception:
                continue

            filas.append({
                "tomo": ws[f"{COL_TOMO}{fila}"].value,
                "anio": ws[f"{COL_ANY}{fila}"].value,
                "matriz_inicio": ws[f"{COL_PROT_INICIAL}{fila}"].value,
                "fecha_inicio": ws[f"{COL_DATA_INICIAL}{fila}"].value,
                "matriz_final": ws[f"{COL_PROT_FINAL}{fila}"].value,
                "fecha_final": ws[f"{COL_DATA_FINAL}{fila}"].value,
                "medida": ws[f"{COL_GRUIX}{fila}"].value,
                "observaciones": ws[f"{COL_OBSERVACIONS}{fila}"].value or "",
            })

        return filas

    # ======================================================
    # INTERFAZ
    # ======================================================

    def _crear_interfaz(self):
        self._crear_cabecera()
        self._crear_selector_area()

        self.area_contenido = ctk.CTkFrame(self, fg_color="transparent")
        self.area_contenido.pack(fill="both", expand=True)

        self.area_medido = ctk.CTkFrame(self.area_contenido, fg_color="transparent")
        self.area_medido.pack(fill="both", expand=True)

        self.area_control = None

        self._crear_barra_acciones()
        self._crear_panel_entrada()
        self._crear_panel_tabla()

        self.matriz_inicio.focus_set()

    def _crear_cabecera(self):
        cabecera = ctk.CTkFrame(self, fg_color=COLOR_PANEL, height=90, corner_radius=0)
        cabecera.pack(fill="x")

        ctk.CTkButton(
            cabecera,
            text="‹ VOLVER",
            command=self.app.mostrar_inicio,
            width=95,
            height=34,
            fg_color="transparent",
            hover_color=COLOR_PANEL_2,
            text_color=COLOR_TEXT,
            font=FONT_NORMAL,
            corner_radius=8,
        ).place(x=22, y=13)

        archivo = str(self.ruta_excel) if self.ruta_excel else "SIN ARCHIVO"

        ctk.CTkLabel(
            cabecera,
            text=f"ARCHIVO: {archivo}",
            font=FONT_SUBTITLE,
            text_color=COLOR_TEXT,
        ).place(x=130, y=17)

        ctk.CTkLabel(
            cabecera,
            text="● GUARDADO",
            font=FONT_NORMAL,
            text_color=COLOR_GREEN,
        ).place(x=130, y=52)

        info = (
            f"TIPO: {self.contexto.tipo or '-'}    "
            f"AÑO: {self.contexto.anio or '-'}\n"
            f"NOTARIO: {self.contexto.notario or '-'}\n"
            f"MEDIDA BASE: {self.medida_estandar}"
        )

        panel_info = ctk.CTkFrame(
            cabecera,
            fg_color=COLOR_PANEL_2,
            border_width=1,
            border_color=COLOR_BORDER,
            corner_radius=10,
            width=380,
            height=68,
        )
        panel_info.place(relx=0.98, y=11, anchor="ne")

        ctk.CTkLabel(
            panel_info,
            text=info,
            font=("Segoe UI", 13, "bold"),
            text_color=COLOR_CYAN,
            justify="left",
        ).place(x=15, y=7)

    def _crear_selector_area(self):
        selector = ctk.CTkFrame(self, fg_color=COLOR_BG, height=48)
        selector.pack(fill="x", padx=25, pady=(12, 0))

        self.boton_area_medido = ctk.CTkButton(
            selector,
            text="MEDIDO",
            command=lambda: self._mostrar_area("medido"),
            width=145,
            height=38,
            fg_color=COLOR_GREEN,
            hover_color=COLOR_GREEN,
            text_color=COLOR_TEXT,
            font=("Segoe UI", 14, "bold"),
            corner_radius=9,
        )
        self.boton_area_medido.pack(side="left")

        self.boton_area_control = ctk.CTkButton(
            selector,
            text="CONTROL",
            command=lambda: self._mostrar_area("control"),
            width=145,
            height=38,
            fg_color=COLOR_PANEL_2,
            hover_color=COLOR_BORDER,
            text_color=COLOR_TEXT,
            font=("Segoe UI", 14, "bold"),
            corner_radius=9,
        )
        self.boton_area_control.pack(side="left", padx=(8, 0))

    def _mostrar_area(self, area):
        if area == "control":
            self.area_medido.pack_forget()

            if self.area_control is None:
                self.area_control = PanelControl(self.area_contenido, self.app)

            self.area_control.pack(fill="both", expand=True)
            self.boton_area_medido.configure(
                fg_color=COLOR_PANEL_2,
                hover_color=COLOR_BORDER,
            )
            self.boton_area_control.configure(
                fg_color=COLOR_GREEN,
                hover_color=COLOR_GREEN,
            )
            return

        if self.area_control is not None:
            self.area_control.pack_forget()

        self.area_medido.pack(fill="both", expand=True)
        self.boton_area_medido.configure(
            fg_color=COLOR_GREEN,
            hover_color=COLOR_GREEN,
        )
        self.boton_area_control.configure(
            fg_color=COLOR_PANEL_2,
            hover_color=COLOR_BORDER,
        )
        self.matriz_inicio.focus_set()

    def _crear_barra_acciones(self):
        barra = ctk.CTkFrame(
            self.area_medido,
            fg_color=COLOR_PANEL,
            border_width=1,
            border_color=COLOR_BORDER,
            corner_radius=12,
            height=58,
        )
        barra.pack(fill="x", padx=25, pady=(12, 0))
        barra.pack_propagate(False)

        ctk.CTkButton(
            barra,
            text="MODIFICAR TOMO",
            command=self._modificar_tomo,
            width=165,
            height=36,
            fg_color=COLOR_PANEL_2,
            hover_color=COLOR_BORDER,
            text_color=COLOR_TEXT,
            font=FONT_NORMAL,
            corner_radius=9,
        ).pack(side="left", padx=(14, 7), pady=10)

        ctk.CTkButton(
            barra,
            text="AÑADIR FOLIADO",
            command=self._foliado,
            width=165,
            height=36,
            fg_color=COLOR_PANEL_2,
            hover_color=COLOR_BORDER,
            text_color=COLOR_TEXT,
            font=FONT_NORMAL,
            corner_radius=9,
        ).pack(side="left", padx=7, pady=10)

        ctk.CTkButton(
            barra,
            text="CERRAR TEMPORADA",
            command=self._cerrar_temporada,
            width=175,
            height=36,
            fg_color=COLOR_PANEL_2,
            hover_color=COLOR_BORDER,
            border_width=1,
            border_color=COLOR_RED,
            text_color=COLOR_TEXT,
            font=FONT_NORMAL,
            corner_radius=9,
        ).pack(side="right", padx=14, pady=10)

    def _crear_panel_entrada(self):
        panel = ctk.CTkFrame(
            self.area_medido,
            fg_color=COLOR_PANEL,
            border_width=1,
            border_color=COLOR_BORDER,
            corner_radius=16,
            height=300,
        )
        panel.pack(fill="x", padx=25, pady=(20, 10))

        ctk.CTkLabel(
            panel,
            text="TOMO ACTUAL",
            font=FONT_NORMAL,
            text_color=COLOR_TEXT_MUTED,
        ).place(x=35, y=30)

        self.label_tomo = ctk.CTkLabel(
            panel,
            text=str(self.tomo_actual),
            font=FONT_TOMO,
            text_color=COLOR_GREEN,
            fg_color=COLOR_PANEL_2,
            width=150,
            height=85,
            corner_radius=14,
        )
        self.label_tomo.place(x=35, y=65)

        ctk.CTkLabel(
            panel,
            text="MATRIZ INICIO ESPERADA",
            font=("Segoe UI", 12, "bold"),
            text_color=COLOR_TEXT_MUTED,
        ).place(x=35, y=165)

        self.label_matriz_esperada = ctk.CTkLabel(
            panel,
            text="LIBRE",
            font=FONT_SUBTITLE,
            text_color=COLOR_CYAN,
        )
        self.label_matriz_esperada.place(x=35, y=190)

        x_label = 235
        x_entry = 430
        y = 32
        salto = 45

        self.matriz_inicio = self._crear_campo(panel, "MATRIZ INICIO", x_label, x_entry, y, ancho=240)
        y += salto
        self.fecha_inicio = self._crear_campo(panel, "FECHA INICIO", x_label, x_entry, y, ancho=240)
        y += salto
        self.matriz_final = self._crear_campo(panel, "MATRIZ FINAL", x_label, x_entry, y, ancho=240)
        y += salto
        self.fecha_final = self._crear_campo(panel, "FECHA FINAL", x_label, x_entry, y, ancho=240)
        y += salto
        self.medida = self._crear_campo(panel, "MEDIDA", x_label, x_entry, y, ancho=140)
        self.medida.insert(0, self.medida_estandar)
        y += salto
        self.observaciones = self._crear_campo(panel, "OBSERVACIONES", x_label, x_entry, y, ancho=620)

        self.mensaje = ctk.CTkLabel(panel, text="", font=FONT_SUBTITLE, text_color=COLOR_GREEN)
        self.mensaje.place(x=35, y=245)

        self._autocompletar_matriz_inicio()
        self._autocompletar_fecha_inicio()

    def _crear_panel_tabla(self):
        panel = ctk.CTkFrame(
            self.area_medido,
            fg_color=COLOR_PANEL,
            border_width=1,
            border_color=COLOR_BORDER,
            corner_radius=16,
        )
        panel.pack(fill="both", expand=True, padx=25, pady=(10, 20))

        cabecera_excel = ctk.CTkFrame(panel, fg_color="transparent", height=48)
        cabecera_excel.pack(fill="x", padx=20, pady=(8, 4))
        cabecera_excel.pack_propagate(False)

        ctk.CTkLabel(
            cabecera_excel,
            text="VISTA COMPLETA DEL EXCEL",
            font=FONT_SUBTITLE,
            text_color=COLOR_TEXT,
        ).pack(side="left", pady=10)

        ctk.CTkButton(
            cabecera_excel,
            text="🔍",
            command=self._buscador,
            width=42,
            height=34,
            fg_color=COLOR_PANEL_2,
            hover_color=COLOR_BORDER,
            text_color=COLOR_TEXT,
            font=("Segoe UI Emoji", 16),
            corner_radius=8,
        ).pack(side="right", pady=7)

        contenedor = ctk.CTkFrame(panel, fg_color=COLOR_PANEL_2, corner_radius=10)
        contenedor.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        columnas = (
            "tomo",
            "any",
            "prot_inicial",
            "data_inicial",
            "prot_final",
            "data_final",
            "gruix",
            "observacions",
        )

        self.tabla = ttk.Treeview(contenedor, columns=columnas, show="headings")

        encabezados = {
            "tomo": "TOMO",
            "any": "ANY",
            "prot_inicial": "PROT INICIAL",
            "data_inicial": "DATA INICIAL",
            "prot_final": "PROT FINAL",
            "data_final": "DATA FINAL",
            "gruix": "GRUIX",
            "observacions": "OBSERVACIONS",
        }

        anchos = {
            "tomo": 70,
            "any": 80,
            "prot_inicial": 130,
            "data_inicial": 130,
            "prot_final": 130,
            "data_final": 130,
            "gruix": 90,
            "observacions": 500,
        }

        for col in columnas:
            self.tabla.heading(col, text=encabezados[col])
            self.tabla.column(col, width=anchos[col], anchor="center")

        self.tabla.column("observacions", anchor="center")

        scroll_y = ttk.Scrollbar(contenedor, orient="vertical", command=self.tabla.yview)
        scroll_x = ttk.Scrollbar(contenedor, orient="horizontal", command=self.tabla.xview)

        self.tabla.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        self.tabla.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")

        contenedor.grid_rowconfigure(0, weight=1)
        contenedor.grid_columnconfigure(0, weight=1)

        self._configurar_estilo_tabla()

    def _crear_campo(self, master, etiqueta, x_label, x_entry, y, ancho):
        ctk.CTkLabel(master, text=etiqueta, font=FONT_NORMAL, text_color=COLOR_TEXT).place(x=x_label, y=y)

        campo = ctk.CTkEntry(
            master,
            width=ancho,
            height=36,
            fg_color=COLOR_PANEL_2,
            border_color=COLOR_BORDER,
            text_color=COLOR_TEXT,
            font=FONT_NORMAL,
            corner_radius=8,
        )
        campo.place(x=x_entry, y=y - 5)

        campo.bind("<KeyRelease>", self._mayusculas)
        campo.bind("<FocusIn>", lambda _event, c=campo: c.configure(border_color=COLOR_CYAN))
        campo.bind("<FocusOut>", lambda _event, c=campo: c.configure(border_color=COLOR_BORDER))

        return campo

    def _boton_secundario(self, master, texto, comando):
        return ctk.CTkButton(
            master,
            text=texto,
            command=comando,
            width=170,
            height=42,
            fg_color=COLOR_PANEL_2,
            hover_color=COLOR_BORDER,
            text_color=COLOR_TEXT,
            font=FONT_NORMAL,
            corner_radius=10,
        )

    def _foliado(self):
        """Abre el formulario independiente de foliado de la temporada."""
        if not self.ruta_excel or not self.ruta_excel.exists():
            messagebox.showerror("Sin temporada", "No hay ningún Excel abierto.")
            return

        VentanaFoliado(
            master=self,
            ruta_excel=self.ruta_excel,
            al_guardar=self._refrescar_tabla_desde_excel,
        )

    def _cerrar_temporada(self):
        """Confirma y genera la línea visual de cierre de la temporada."""
        if not self.ruta_excel or not self.ruta_excel.exists():
            messagebox.showerror("Sin temporada", "No hay ningún Excel abierto.")
            return

        if self.tomo_actual == 1:
            messagebox.showwarning(
                "No se puede cerrar la temporada",
                "Todavía no existen tomos medidos.",
            )
            return

        if temporada_cerrada(self.ruta_excel):
            messagebox.showwarning(
                "Temporada ya cerrada",
                "Esta temporada ya contiene una línea de cierre.\n\nNo se realizarán cambios.",
            )
            return

        confirmar = messagebox.askyesno(
            "Cerrar temporada",
            "Se marcará como finalizado el medido de esta temporada.\n\n¿Deseas continuar?",
            parent=self,
        )
        if not confirmar:
            return

        try:
            cerrar_temporada(self.ruta_excel, self.contexto.notario)
        except ValueError:
            messagebox.showwarning(
                "No se puede cerrar la temporada",
                "Todavía no existen tomos medidos.",
            )
            return
        except RuntimeError:
            messagebox.showwarning(
                "Temporada ya cerrada",
                "Esta temporada ya contiene una línea de cierre.\n\nNo se realizarán cambios.",
            )
            return
        except Exception as error:
            messagebox.showerror(
                "Error al cerrar temporada",
                f"No se pudo guardar el cierre de la temporada:\n\n{error}",
            )
            return

        messagebox.showinfo(
            "Temporada cerrada",
            "La temporada se ha cerrado correctamente.",
        )

    def _modificar_tomo(self):
        """Abre el formulario de corrección de tomos anteriores."""
        if not self.ruta_excel or not self.ruta_excel.exists():
            messagebox.showerror("Sin temporada", "No hay ningún Excel abierto.")
            return

        VentanaModificarTomo(
            master=self,
            ruta_excel=self.ruta_excel,
            medida_estandar=self.medida_estandar,
            al_guardar=self._recargar_despues_de_modificar,
        )

    def _recargar_despues_de_modificar(self):
        self._cargar_estado_excel()
        self._actualizar_tomo()
        self._actualizar_matriz_esperada()
        self._refrescar_tabla_desde_excel()

    # ======================================================
    # TABLA
    # ======================================================

    def _configurar_estilo_tabla(self):
        style = ttk.Style()
        style.theme_use("default")

        style.configure(
            "Treeview",
            background="#181818",
            foreground=COLOR_TEXT,
            fieldbackground="#181818",
            rowheight=28,
            font=("Segoe UI", 12),
        )

        style.configure(
            "Treeview.Heading",
            background=COLOR_PANEL_2,
            foreground=COLOR_CYAN,
            font=("Segoe UI", 12, "bold"),
        )

        style.map("Treeview", background=[("selected", "#003C5A")])

        self.tabla.tag_configure("normal", background="#181818", foreground=COLOR_TEXT)
        self.tabla.tag_configure("especial", background=COLOR_RED, foreground="#FFFFFF")
        self.tabla.tag_configure("interrogante", background=COLOR_YELLOW, foreground="#000000")

    def _refrescar_tabla_desde_excel(self):
        for item in self.tabla.get_children():
            self.tabla.delete(item)

        for fila in self._leer_filas_excel():
            medida = str(fila["medida"] or "")
            tag = self._tag_medida(medida)

            item = self.tabla.insert(
                "",
                "end",
                values=(
                    fila["tomo"],
                    fila["anio"],
                    fila["matriz_inicio"],
                    fila["fecha_inicio"],
                    fila["matriz_final"],
                    fila["fecha_final"],
                    fila["medida"],
                    fila["observaciones"],
                ),
                tags=(tag,),
            )
            self.tabla.see(item)

    # ======================================================
    # TECLADO Y VALIDACIONES
    # ======================================================

    def _configurar_enter(self):
        campos = [
            self.matriz_inicio,
            self.fecha_inicio,
            self.matriz_final,
            self.fecha_final,
            self.medida,
            self.observaciones,
        ]

        for i, campo in enumerate(campos):
            if i < len(campos) - 1:
                campo.bind("<Return>", lambda _event, idx=i: self._enter_campo(campos, idx))
            else:
                campo.bind("<Return>", lambda _event: self._guardar_tomo())

        self.fecha_inicio.bind("<FocusOut>", lambda _event: self._normalizar_mes(self.fecha_inicio))
        self.fecha_final.bind("<FocusOut>", lambda _event: self._normalizar_mes(self.fecha_final))

    def _enter_campo(self, campos, idx):
        campo_actual = campos[idx]

        if campo_actual in (self.fecha_inicio, self.fecha_final):
            self._normalizar_mes(campo_actual)

        campos[idx + 1].focus_set()
        return "break"

    def _normalizar_mes(self, campo):
        """
        Normaliza fechas escritas como día + mes.

        Ejemplos:
        1 EN   -> 1 ENERO
        15 FEB -> 15 FEBRERO
        7 S    -> 7 SEPTIEMBRE

        También acepta solo mes:
        EN -> ENERO

        No modifica formatos con barras, como 12/01/2025.
        """

        valor_original = campo.get().strip()
        valor = valor_original.upper()

        if not valor:
            return

        if "/" in valor or "-" in valor:
            return

        partes = valor.split()

        if len(partes) == 1:
            if partes[0].isalpha():
                mes = MESES_AUTOCOMPLETAR.get(partes[0])
                if mes:
                    campo.delete(0, "end")
                    campo.insert(0, mes)
            return

        if len(partes) < 2:
            return

        dia_texto = partes[0]
        mes_texto = partes[1]

        if not dia_texto.isdigit():
            return

        mes = MESES_AUTOCOMPLETAR.get(mes_texto)

        if not mes:
            return

        dia = int(dia_texto)

        campo.delete(0, "end")
        campo.insert(0, f"{dia} {mes}")

    def _mayusculas(self, event):
        campo = event.widget
        texto = campo.get()
        cursor = campo.index("insert")

        campo.delete(0, "end")
        campo.insert(0, texto.upper())
        campo.icursor(cursor)

    def _validar_fecha_texto(self, valor):
        """
        Valida fechas tipo: 1 ENERO, 15 FEBRERO, 31 DICIEMBRE.

        Si el formato no es día + mes, no bloquea, porque de momento
        aceptamos texto libre y fechas manuales.
        """

        valor = valor.strip().upper()

        if not valor:
            return False, "Fecha vacía."

        if "/" in valor or "-" in valor:
            return True, ""

        partes = valor.split()

        if len(partes) != 2:
            return True, ""

        dia_texto, mes = partes

        if not dia_texto.isdigit():
            return True, ""

        if mes not in DIAS_MAXIMOS_MES:
            return True, ""

        dia = int(dia_texto)
        maximo = DIAS_MAXIMOS_MES[mes]

        if dia < 1 or dia > maximo:
            return False, f"Día incorrecto para {mes}. Máximo: {maximo}."

        return True, ""

    def _validar(self):
        if not self.matriz_inicio.get().strip():
            return False, "Falta matriz inicio.", self.matriz_inicio

        if not self.fecha_inicio.get().strip():
            return False, "Falta fecha inicio.", self.fecha_inicio

        if not self.matriz_final.get().strip():
            return False, "Falta matriz final.", self.matriz_final

        if not self.fecha_final.get().strip():
            return False, "Falta fecha final.", self.fecha_final

        self._normalizar_mes(self.fecha_inicio)
        self._normalizar_mes(self.fecha_final)

        fecha_ok, fecha_msg = self._validar_fecha_texto(self.fecha_inicio.get())
        if not fecha_ok:
            return False, fecha_msg, self.fecha_inicio

        fecha_ok, fecha_msg = self._validar_fecha_texto(self.fecha_final.get())
        if not fecha_ok:
            return False, fecha_msg, self.fecha_final

        if not self.medida.get().strip():
            return False, "Falta medida.", self.medida

        if not self.matriz_inicio.get().strip().isdigit():
            return False, "Matriz inicio debe ser numérica.", self.matriz_inicio

        if not self.matriz_final.get().strip().isdigit():
            return False, "Matriz final debe ser numérica.", self.matriz_final

        matriz_inicio = int(self.matriz_inicio.get().strip())
        matriz_final = int(self.matriz_final.get().strip())

        if matriz_final < matriz_inicio:
            return False, "Protocolo final no puede ser menor que protocolo inicial.", self.matriz_final

        if self.tomo_actual > 1 and self.matriz_inicio_esperada is not None:
            if matriz_inicio != self.matriz_inicio_esperada:
                return False, f"Matriz inicio incorrecta. Esperada: {self.matriz_inicio_esperada}.", self.matriz_inicio

        medida = self.medida.get().strip()

        if not self._validar_medida(medida):
            return False, "Medida incorrecta. Usa número hasta 10 con coma, o ?.", self.medida

        if len(self.observaciones.get()) > MAX_OBSERVACIONES:
            return False, f"Observaciones no puede superar {MAX_OBSERVACIONES} caracteres.", self.observaciones

        return True, "", None

    def _validar_medida(self, medida):
        if medida == "?":
            return True

        try:
            valor = float(medida.replace(",", "."))
        except ValueError:
            return False

        return 0 < valor <= MAX_MEDIDA

    # ======================================================
    # GUARDADO REAL
    # ======================================================

    def _guardar_tomo(self):
        if self.ruta_excel and self.ruta_excel.exists() and temporada_cerrada(self.ruta_excel):
            messagebox.showwarning(
                "Temporada ya cerrada",
                "Esta temporada ya contiene una línea de cierre.\n\nNo se pueden añadir nuevos tomos mientras exista el cierre.",
            )
            return
        ok, mensaje, campo = self._validar()

        if not ok:
            messagebox.showwarning("Datos incorrectos", mensaje)
            if campo:
                campo.focus_set()
            return

        datos = {
            "tomo": self.tomo_actual,
            "anio": self.contexto.anio or "",
            "matriz_inicio": self.matriz_inicio.get().strip(),
            "fecha_inicio": self.fecha_inicio.get().strip(),
            "matriz_final": self.matriz_final.get().strip(),
            "fecha_final": self.fecha_final.get().strip(),
            "medida": self.medida.get().strip(),
            "observaciones": self.observaciones.get().strip(),
        }

        try:
            self._guardar_en_excel(datos)
        except Exception as e:
            messagebox.showerror("Error al guardar", f"No se pudo guardar en Excel:\n\n{e}")
            return

        matriz_final_guardada = int(datos["matriz_final"])
        self.ultima_matriz_final = matriz_final_guardada
        self.ultima_fecha_final = datos["fecha_final"]
        self.matriz_inicio_esperada = matriz_final_guardada + 1

        self._cargar_estado_excel()
        self._refrescar_tabla_desde_excel()

        self.mensaje.configure(text=f"✔ TOMO {datos['tomo']} GUARDADO")
        self.after(1000, lambda: self.mensaje.configure(text=""))

        self._limpiar_formulario()
        self._actualizar_tomo()
        self._actualizar_matriz_esperada()
        self._autocompletar_matriz_inicio()
        self._autocompletar_fecha_inicio()

        self.fecha_inicio.focus_set()

    def _tag_medida(self, medida):
        if medida == "?":
            return "interrogante"

        if medida != self.medida_estandar:
            return "especial"

        return "normal"

    def _limpiar_formulario(self):
        self.matriz_inicio.delete(0, "end")
        self.fecha_inicio.delete(0, "end")
        self.matriz_final.delete(0, "end")
        self.fecha_final.delete(0, "end")
        self.medida.delete(0, "end")
        self.observaciones.delete(0, "end")

        self.medida.insert(0, self.medida_estandar)

    def _actualizar_tomo(self):
        self.label_tomo.configure(text=str(self.tomo_actual))

    def _actualizar_matriz_esperada(self):
        if self.matriz_inicio_esperada is None:
            self.label_matriz_esperada.configure(text="LIBRE", text_color=COLOR_CYAN)
        else:
            self.label_matriz_esperada.configure(text=str(self.matriz_inicio_esperada), text_color=COLOR_GREEN)

    def _autocompletar_matriz_inicio(self):
        if self.matriz_inicio_esperada is not None:
            self.matriz_inicio.delete(0, "end")
            self.matriz_inicio.insert(0, str(self.matriz_inicio_esperada))

    def _autocompletar_fecha_inicio(self):
        """
        Copia la fecha final del tomo anterior como fecha inicial del siguiente.
        El campo queda editable.
        """

        if self.ultima_fecha_final:
            self.fecha_inicio.delete(0, "end")
            self.fecha_inicio.insert(0, str(self.ultima_fecha_final))

    # ======================================================
    # BOTONES
    # ======================================================

    def _buscador(self):
        messagebox.showinfo("Buscador", "Buscador pendiente.")

